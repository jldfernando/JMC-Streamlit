### Function scripts for JMC Streamlit Application

## Libraries
import pandas as pd
import numpy as np
import streamlit as st
from datetime import datetime, time
import openpyxl

##Variables:
jmc_group_filters = {
    'Capts': ['Therese Reyes', 'Gabby Castillo', 'Josh Fernando'],
    'MP': ['Andie Miranda'],
    'SU Heads': ['James Calilung','JC Legaspi','Princess Mangubat','Marcus Tornea'],
    '15th Fleet': ['Eara Cayañga','Erika Jarder','Andie Miranda','Angel N.','Dani Peralta','Margaux Perez','Therese Reyes'],
    '16th Fleet': ['MJ Abeleda','Coco Ballonado','James Calilung','Gabby Castillo','Matthew Caudal','EJ De Lima','Josh Fernando',
                   'Theresa Imperial','JC Legaspi','Princess Mangubat','Marcus Tornea'],
    '17th Fleet': ['Juliana Bautista', 'Prudence Clemente', 'Ghia Espino', 'Seth Fuentes', 'Carlo Galura']
}
all_jmcs = jmc_group_filters['15th Fleet'] + jmc_group_filters ['16th Fleet'] + jmc_group_filters['17th Fleet']
jmc_filters = ['Eara Cayañga','Erika Jarder','Andie Miranda','Angel Nicole','Dani Peralta','Margaux Perez','Therese Reyes',
               'MJ Abeleda','Coco Ballonado','Gabby Castillo','Matthew Caudal','EJ De Lima','Josh Fernando',
               'Theresa Imperial','JC Legaspi','Princess Mangubat','Marcus Tornea',
               'Juliana Bautista', 'Prudence Clemente', 'Ghia Espino', 'Seth Fuentes', 'Carlo Galura']

event_options = ['Visit','Tour','Event']

## FUNCTIONS

# Phased out Function
def get_worksheet_names(file) -> list:
    """
    PHASED OUT
    Input: excel file address
    Output: worksheets in excel file
    """
    try:
        workbook = openpyxl.load_workbook(file)
        return workbook.sheetnames

    except FileNotFoundError:
        print(f"Error: File '{file}' not found.")
        return []
    except Exception as e:
        print(f"An error occurred: {e}")
        return []

def get_indiv_schedules_dict(file:str, sheets_to_drop: list = None) -> dict:
    """
    Get a dictionary of the schedules stored in a dataframe with the key being the worksheet name
    Input: file address, list of worksheets to drop
    Output: dictionary of schedules as dataframes
    Note: format of excel file is very specific, see details in TBA
    """
    sheetnames = openpyxl.load_workbook(file).sheetnames
    indiv_schedules = {}
    
    for sheet in sheets_to_drop:
        if sheet in sheetnames:
            sheetnames.remove(sheet)
    
    for sheet in sheetnames:
        df = pd.read_excel(file, sheet, skiprows=1) #assuming that the first row is the name
        df.set_index('TIME', inplace=True)
        indiv_schedules.update({sheet: df})
        
    return indiv_schedules
    
def compile_schedules(schedules_dict: pd.DataFrame) -> pd.DataFrame:
    """
    Convert schedules into a more dataset-like format and merge them into one dataframe
    Input: A dictionary of schedule dataframes
    Output: A dataframe with all schedules
    Note: format of schedule is very specific, see details in TBA
    """
    temp_scheds = []
    
    for schedule in schedules_dict:
        df = schedules_dict[schedule]
        df.reset_index(drop=False, inplace=True)
        df['hours'] = np.linspace(7.5,20.5,27)
        df_melted = df.melt(id_vars=['TIME','hours'], var_name='day', value_name='available')
        df_melted['available'] = df_melted['available'].isna()
        df_melted['name'] = schedule # Use schedule.split()[0] to get only the firstnames
        df_melted.rename(columns={'TIME':'time'}, inplace=True)
        temp_scheds.append(df_melted)
        
    df_compiled = pd.concat([df for df in temp_scheds], axis=0)
    return df_compiled

def df_highlight_value(df: pd.DataFrame, highlights=['x', 'X'], color='#a70012'):
    """
    Highlights specific values in a dataframe
    Input: df, list of values to highlight, color to higlight with
    Output: highlighted df
    """
    def highlight(val):
        if val in highlights:
            return f'background-color: {color}'
        else:
            return ''
    styled_df = df.style.map(highlight).format(precision=0)
    return styled_df

def df_colormap_values(df: pd.DataFrame, cmap='Reds') -> pd.DataFrame:
    """
    Applies a colormap to a dataframe
    Input: df, colormap
    Output: df
    Note: need to update kwargs for df.style.background_gradient
    """
    styled_df = df.style.background_gradient(cmap=cmap).format(precision=0)
    return styled_df

def uncompress_schedule_total_availablity(schedule: pd.DataFrame, people_filter=[], show_names = False) -> pd.DataFrame:
    """
    Given a compiled schedule & people to filter, it will return a dataframe with the total number of those people availiable in schedule-like format
    Input: compiled schedule df, list of people included
    Output: schedule df
    """
    if people_filter != []:
        df_filtered = schedule[schedule['name'].isin(people_filter)]
    else:
        df_filtered = schedule
    
    df_available_only = df_filtered[df_filtered['available'] == True]
    
    df_grouped = df_available_only.groupby(['time','hours','day'])['name'].apply(lambda x: ', '.join(x)).reset_index()
    
    if show_names:
        df_grouped['total_availability'] = df_grouped['name'].apply(lambda x: str(len(x.split(','))))
        df_grouped['named_availability'] = '(' + df_grouped['total_availability'] + ') ' + df_grouped['name']
        df_pivot = df_grouped.pivot(index=['time','hours'], columns='day', values='named_availability')
    else:
        df_grouped['total_availability'] = df_grouped['name'].apply(lambda x: len(x.split(',')))
        df_pivot = df_grouped.pivot(index=['time','hours'], columns='day', values='total_availability')
    
    df_reordered = df_pivot.loc[:, ['MONDAY', 'TUESDAY', 'WEDNESDAY','THURSDAY','FRIDAY','SATURDAY']]
    df_reordered = df_reordered.reset_index().sort_values(by='hours')
    df_reordered.set_index('time', inplace=True)
    df_reordered.drop('hours', axis=1, inplace=True)
    return df_reordered

def update_total_availability_display(schedule: pd.DataFrame, group_filter, people_filter, show_names, exclude):
    """
    Used to update thet dispay given the filters selected
    """
    if group_filter == []:
        total_group = all_jmcs
    else:
        total_group = []
        for group in group_filter:
            for person in jmc_group_filters[group]:
                if person not in total_group:
                    total_group.append(person)
    
    if people_filter == []:
        final_filter = total_group
    else:                
        final_filter = [person for person in total_group if person in people_filter]
        
    if exclude:
        final_filter = [person for person in total_group if person not in final_filter]
        
    updated_display = uncompress_schedule_total_availablity(schedule, final_filter, show_names)
    return updated_display, total_group

def hours_past_midnight(dt: datetime):
    hour = dt.hour
    mins = dt.minute/60
    return hour+mins

def check_availability(event_type, event_date, event_start, event_end, df: pd.DataFrame):
    day_index = event_date.weekday()
    days_of_week = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY", "SUNDAY"]
    filter_day = days_of_week[day_index]
    filter_start = hours_past_midnight(event_start)
    filter_end = hours_past_midnight(event_end)
    if filter_start%0.5 != 0:
        filter_start = filter_start - (filter_start%0.5)
    if filter_end%0.5 != 0:
        filter_end = filter_end - (filter_end%0.5)
    intervals = int(1+(filter_end-filter_start)/0.5)
    filter_hours = np.linspace(filter_start, filter_end, intervals)
    
    filtered_df = df[df['day'] == filter_day]
    filtered_df = filtered_df[filtered_df['hours'].isin(filter_hours)]

    results_available = filtered_df.groupby('name')['available'].sum()
    results_cut = results_available.apply(lambda x: (intervals-x)/2)
    results_cut = results_cut.sort_values()
    results_df = results_cut.to_frame()
    results_df.reset_index(inplace=True)
    results = results_df.groupby('available')['name'].apply(lambda x: ', '.join(x)).reset_index()
    results.rename(columns={'available':'Cuts (hours)', 'name':'JMCs'}, inplace=True)
    results.set_index('Cuts (hours)', inplace=True)
    return results

def format_input(event_name, event_type, event_date, event_start, event_end):
    day = event_date.weekday()
    date = event_date.strftime("%B %d, %A")
    start = event_start.strftime("%I:%M %p")
    end = event_end.strftime("%I:%M %p")
    duration = f'{start} - {end}'
    line = '-'*50
    
    readback = f'{event_type} Details:\
        \n {line}\
        \n\t {event_name}\
        \n\t Date: {date}\
        \n\t Duration: {duration}'
    return readback

def update_indiv_display(jmc_name):
    indiv_df = st.session_state.indiv_schedules[jmc_name]
    display_df = indiv_df.set_index('TIME')
    display_df.drop('hours', axis=1, inplace=True)
    return display_df

def get_visits(time_filter, jmc_filter):
    visits_df = pd.read_csv(r'data/visit_data.csv', index_col=0, parse_dates=['Date'])
    
    if time_filter == '1st Sem':
        visits_df = visits_df[visits_df['Date'].apply(lambda x: x.year) == 2024]
    elif time_filter == '2nd Sem':
        visits_df = visits_df[visits_df['Date'].apply(lambda x: x.year) == 2025]
    
    if jmc_filter == None:
        visits_df = visits_df[~visits_df[['INSTITUTION','Date']].duplicated()]
    else:
        visits_df = visits_df[visits_df['JMCs Assigned'] == jmc_filter]
    
    visits_df.drop(columns=['JMCs Assigned', 'Zoom', 'Gmeet'], inplace=True)
    visits_df['Visit Date'] = visits_df['Date'].apply(lambda x: x.strftime("%b %d, %Y"))
    return visits_df[['INSTITUTION','Visit Date','Day','Onsite','Online','MT','Module','Fair', 'Date']]

def get_tours(time_filter, jmc_filter):
    tours_df = pd.read_csv(r'data/tour_data.csv', index_col=0, parse_dates=['Date'])
    
    if time_filter == '1st Sem':
        tours_df = tours_df[tours_df['Date'].apply(lambda x: x.year) == 2024]
    elif time_filter == '2nd Sem':
        tours_df = tours_df[tours_df['Date'].apply(lambda x: x.year) == 2025]
    
    if jmc_filter == None:
        tours_df = tours_df[~tours_df[['Full Name/School/Unit','Date']].duplicated()]
    else:
        tours_df = tours_df[tours_df['Assigned JMC'] == jmc_filter]
    
    tours_df.drop(columns=['Assigned JMC'], inplace=True)
    tours_df['Tour Date'] = tours_df['Date'].apply(lambda x: x.strftime("%B %d, %Y"))
    return tours_df[['Full Name/School/Unit','Tour Date','Day', 'Date']]

def get_plot_data(data1, data2):
    df1 = data1.groupby('Date').size().sort_index().cumsum()
    df2 = data2.groupby('Date').size().sort_index().cumsum()
    data = pd.concat([df1, df2], axis=1)
    data.columns = ['Visits','Tours']
    data.ffill(inplace=True)
    return data
    